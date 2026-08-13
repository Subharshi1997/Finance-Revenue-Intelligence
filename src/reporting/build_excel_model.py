"""Phase 11: builds the Excel finance operations workbook
(excel/finance_ops_model.xlsx) from a coherent subset of the data - 40
merchants (stratified by segment) with every invoice, payment, credit note,
and collection activity they have, so the XLOOKUP/SUMIFS/COUNTIFS formulas
join correctly across sheets. Raw Transactions is capped at 2,500 rows
(transactions aren't joined row-by-row elsewhere, so subsampling them doesn't
break any other sheet).

Sheets: Merchants (reference), Raw Transactions, Invoices, Payments,
Invoice Validation, AR Aging, Collections, Revenue Reconciliation,
Monthly MIS, KPI Dashboard.

Run from the project root (after the standard data-generation + SQL build):
    python -m src.reporting.build_excel_model
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[2]
PROCESSED_DIR = ROOT / "data" / "processed"
EXCEL_DIR = ROOT / "excel"
OUTPUT_PATH = EXCEL_DIR / "finance_ops_model.xlsx"

RNG_SEED = 42
MERCHANTS_PER_SEGMENT = {"Enterprise": 4, "Mid-Market": 8, "SMB": 14, "Long-tail": 14}
MAX_TRANSACTION_ROWS = 2500

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")


def _select_merchants(merchants: pd.DataFrame) -> pd.DataFrame:
    rng = np.random.default_rng(RNG_SEED)
    parts = []
    for segment, n in MERCHANTS_PER_SEGMENT.items():
        pool = merchants[merchants["merchant_segment"] == segment]
        n = min(n, len(pool))
        idx = rng.choice(pool.index, size=n, replace=False)
        parts.append(pool.loc[idx])
    return pd.concat(parts).sort_values("merchant_id").reset_index(drop=True)


def _coerce_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Converts any column whose name contains 'date' (but not a timestamp)
    to a real Python date object, so Excel stores it as a date value that
    DATE()/subtraction formulas can operate on - not a text string."""
    df = df.copy()
    for col in df.columns:
        if "date" in col.lower() and "timestamp" not in col.lower():
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
    return df


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, table_name: str | None = None) -> int:
    """Writes a DataFrame as a formatted Excel Table starting at start_row.
    Returns the last data row number written."""
    df = _coerce_date_columns(df)
    for c_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    date_cols = {c_idx for c_idx, col in enumerate(df.columns, start=1) if "date" in col.lower() and "timestamp" not in col.lower()}
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, (np.integer,)):
                value = int(value)
            elif isinstance(value, (np.floating,)):
                value = float(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx in date_cols and value is not None:
                cell.number_format = "yyyy-mm-dd"

    last_row = start_row + len(df)
    last_col_letter = get_column_letter(len(df.columns))
    if table_name:
        ref = f"A{start_row}:{last_col_letter}{last_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(table)
    for c_idx, col in enumerate(df.columns, start=1):
        width = max(11, min(28, int(df[col].astype(str).str.len().quantile(0.9)) + 2))
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.freeze_panes = f"A{start_row + 1}"
    return last_row


def build_merchants_sheet(wb, merchants: pd.DataFrame) -> None:
    ws = wb.create_sheet("Merchants")
    ws["A1"] = "Merchant Reference (lookup source for XLOOKUP formulas on other sheets)"
    ws["A1"].font = TITLE_FONT
    _write_df(ws, merchants, start_row=3, table_name="tblMerchants")


def build_raw_transactions_sheet(wb, transactions: pd.DataFrame) -> None:
    ws = wb.create_sheet("Raw Transactions")
    ws["A1"] = f"Raw Transactions (sampled: {len(transactions):,} of the full dataset for these merchants)"
    ws["A1"].font = TITLE_FONT
    _write_df(ws, transactions, start_row=3, table_name="tblTransactions")


def build_invoices_sheet(wb, invoices: pd.DataFrame) -> int:
    ws = wb.create_sheet("Invoices")
    ws["A1"] = "Invoices"
    ws["A1"].font = TITLE_FONT
    last_row = _write_df(ws, invoices, start_row=3, table_name="tblInvoices")
    return last_row


def build_payments_sheet(wb, payments: pd.DataFrame) -> None:
    ws = wb.create_sheet("Payments")
    ws["A1"] = "Payments"
    ws["A1"].font = TITLE_FONT
    _write_df(ws, payments, start_row=3, table_name="tblPayments")


def build_invoice_validation_sheet(wb, invoices: pd.DataFrame) -> None:
    """Merchant name via XLOOKUP against the Merchants sheet; billing status
    via nested IF comparing billed_fee to expected_fee."""
    ws = wb.create_sheet("Invoice Validation")
    ws["A1"] = "Invoice Validation - expected vs. billed, with XLOOKUP + IF formulas"
    ws["A1"].font = TITLE_FONT

    header = [
        "invoice_id", "merchant_id", "merchant_name", "expected_fee", "billed_fee",
        "revenue_difference", "variance_percent", "billing_status", "review_status",
    ]
    start_row = 3
    for c_idx, col in enumerate(header, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    n = len(invoices)
    for i, row in enumerate(invoices.itertuples(index=False), start=1):
        r = start_row + i
        ws.cell(row=r, column=1, value=row.invoice_id)
        ws.cell(row=r, column=2, value=row.merchant_id)
        # XLOOKUP into the Merchants reference sheet
        ws.cell(row=r, column=3, value=(
            f'=XLOOKUP(B{r},Merchants!$A$4:$A$1000,Merchants!$B$4:$B$1000,"Not found")'
        ))
        ws.cell(row=r, column=4, value=(
            f'=XLOOKUP(A{r},tblInvoices[invoice_id],tblInvoices[expected_fee])'
        ))
        ws.cell(row=r, column=5, value=(
            f'=XLOOKUP(A{r},tblInvoices[invoice_id],tblInvoices[billed_fee])'
        ))
        ws.cell(row=r, column=6, value=f"=E{r}-D{r}")
        ws.cell(row=r, column=7, value=f'=IF(D{r}=0,0,ROUND(100*F{r}/D{r},2))')
        ws.cell(row=r, column=8, value=(
            f'=IF(ABS(F{r})<=0.01,"Correct",IF(F{r}<0,"Underbilled","Overbilled"))'
        ))

    last_row = start_row + n
    last_col_letter = get_column_letter(len(header))
    table = Table(displayName="tblInvoiceValidation", ref=f"A{start_row}:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    # Data Validation: manual reviewer sign-off dropdown, left blank for the
    # analyst to fill in - billing_status is system-derived, review_status
    # is the human control step on top of it.
    review_dv = DataValidation(
        type="list",
        formula1='"Pending Review,Reviewed - No Action,Escalated to Billing,Corrected"',
        allow_blank=True,
    )
    review_dv.error = "Choose a value from the dropdown list."
    review_dv.errorTitle = "Invalid review status"
    ws.add_data_validation(review_dv)
    review_dv.add(f"I{start_row + 1}:I{last_row}")

    ws.conditional_formatting.add(
        f"H{start_row + 1}:H{last_row}",
        CellIsRule(operator="equal", formula=['"Underbilled"'], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    ws.conditional_formatting.add(
        f"H{start_row + 1}:H{last_row}",
        CellIsRule(operator="equal", formula=['"Overbilled"'], fill=PatternFill("solid", fgColor="FFEB9C")),
    )
    ws.conditional_formatting.add(
        f"H{start_row + 1}:H{last_row}",
        CellIsRule(operator="equal", formula=['"Correct"'], fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    for c_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 16
    ws.freeze_panes = f"A{start_row + 1}"


def build_ar_aging_sheet(wb, invoices: pd.DataFrame, as_of_date: str) -> None:
    """SUMIFS against Payments for total_paid, COUNTIFS-style aging via IFS."""
    ws = wb.create_sheet("AR Aging")
    ws["A1"] = "AR Aging - SUMIFS against Payments, aging bucket via IFS"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"As-of date: {as_of_date}"

    header = [
        "invoice_id", "merchant_id", "merchant_name", "due_date", "total_invoice_amount",
        "total_paid", "outstanding_amount", "days_overdue", "aging_bucket",
    ]
    start_row = 4
    for c_idx, col in enumerate(header, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    n = len(invoices)
    for i, row in enumerate(invoices.itertuples(index=False), start=1):
        r = start_row + i
        ws.cell(row=r, column=1, value=row.invoice_id)
        ws.cell(row=r, column=2, value=row.merchant_id)
        ws.cell(row=r, column=3, value=f'=XLOOKUP(B{r},Merchants!$A$4:$A$1000,Merchants!$B$4:$B$1000,"Not found")')
        ws.cell(row=r, column=4, value=f'=XLOOKUP(A{r},tblInvoices[invoice_id],tblInvoices[due_date])')
        ws.cell(row=r, column=5, value=f'=XLOOKUP(A{r},tblInvoices[invoice_id],tblInvoices[total_invoice_amount])')
        ws.cell(row=r, column=6, value=(
            f'=SUMIFS(tblPayments[payment_amount],tblPayments[invoice_id],A{r},'
            f'tblPayments[payment_status],"Success")'
        ))
        ws.cell(row=r, column=7, value=f"=MAX(0,E{r}-F{r})")
        ws.cell(row=r, column=8, value=f'=DATE({as_of_date[:4]},{int(as_of_date[5:7])},{int(as_of_date[8:10])})-D{r}')
        ws.cell(row=r, column=9, value=(
            f'=IFS(G{r}=0,"Excluded (paid)",H{r}<=0,"Current",H{r}<=30,"1-30 days",'
            f'H{r}<=60,"31-60 days",H{r}<=90,"61-90 days",TRUE,"90+ days")'
        ))

    last_row = start_row + n
    last_col_letter = get_column_letter(len(header))
    table = Table(displayName="tblARAging", ref=f"A{start_row}:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    ws.conditional_formatting.add(
        f"G{start_row + 1}:G{last_row}",
        ColorScaleRule(start_type="min", start_color="C6EFCE", end_type="max", end_color="FFC7CE"),
    )
    for c_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 16
    ws.freeze_panes = f"A{start_row + 1}"

    # Aging summary block (SUMIFS/COUNTIFS pivot-style)
    summary_start = start_row + n + 3
    ws.cell(row=summary_start, column=1, value="Aging Summary").font = TITLE_FONT
    buckets = ["Current", "1-30 days", "31-60 days", "61-90 days", "90+ days"]
    ws.cell(row=summary_start + 1, column=1, value="Bucket")
    ws.cell(row=summary_start + 1, column=2, value="Invoice Count")
    ws.cell(row=summary_start + 1, column=3, value="Total Outstanding")
    for i, bucket in enumerate(buckets, start=1):
        r = summary_start + 1 + i
        ws.cell(row=r, column=1, value=bucket)
        ws.cell(row=r, column=2, value=f'=COUNTIFS(I{start_row + 1}:I{last_row},A{r})')
        ws.cell(row=r, column=3, value=f'=SUMIFS(G{start_row + 1}:G{last_row},I{start_row + 1}:I{last_row},A{r})')


def build_collections_sheet(wb, collection_activity: pd.DataFrame) -> None:
    ws = wb.create_sheet("Collections")
    ws["A1"] = "Collection Activity - with outcome-effectiveness SUMIFS/COUNTIFS summary"
    ws["A1"].font = TITLE_FONT
    last_row = _write_df(ws, collection_activity, start_row=3, table_name="tblCollectionActivity")

    summary_start = last_row + 3
    ws.cell(row=summary_start, column=1, value="Channel Effectiveness").font = TITLE_FONT
    ws.cell(row=summary_start + 1, column=1, value="Channel")
    ws.cell(row=summary_start + 1, column=2, value="Total Activities")
    ws.cell(row=summary_start + 1, column=3, value="Payments Received")
    ws.cell(row=summary_start + 1, column=4, value="Success Rate %")
    channels = sorted(collection_activity["contact_channel"].dropna().unique().tolist())
    data_start, data_end = 4, last_row
    for i, channel in enumerate(channels, start=1):
        r = summary_start + 1 + i
        ws.cell(row=r, column=1, value=channel)
        ws.cell(row=r, column=2, value=f'=COUNTIFS(F{data_start}:F{data_end},A{r})')
        ws.cell(row=r, column=3, value=f'=COUNTIFS(F{data_start}:F{data_end},A{r},G{data_start}:G{data_end},"Payment Received")')
        ws.cell(row=r, column=4, value=f'=IF(B{r}=0,0,ROUND(100*C{r}/B{r},2))')


def build_revenue_reconciliation_sheet(wb, invoices: pd.DataFrame) -> None:
    ws = wb.create_sheet("Revenue Reconciliation")
    ws["A1"] = "Revenue Reconciliation - invoice vs. total payments, status via IFS"
    ws["A1"].font = TITLE_FONT

    header = ["invoice_id", "merchant_id", "total_invoice_amount", "total_paid", "variance", "reconciliation_status"]
    start_row = 3
    for c_idx, col in enumerate(header, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    n = len(invoices)
    for i, row in enumerate(invoices.itertuples(index=False), start=1):
        r = start_row + i
        ws.cell(row=r, column=1, value=row.invoice_id)
        ws.cell(row=r, column=2, value=row.merchant_id)
        ws.cell(row=r, column=3, value=f'=XLOOKUP(A{r},tblInvoices[invoice_id],tblInvoices[total_invoice_amount])')
        ws.cell(row=r, column=4, value=(
            f'=SUMIFS(tblPayments[payment_amount],tblPayments[invoice_id],A{r},'
            f'tblPayments[payment_status],"Success")'
        ))
        ws.cell(row=r, column=5, value=f"=D{r}-C{r}")
        ws.cell(row=r, column=6, value=(
            f'=IFS(D{r}=0,"MISSING_PAYMENT",ABS(E{r})<=1,"MATCHED",E{r}>1,"AMOUNT_MISMATCH",TRUE,"PARTIAL")'
        ))

    last_row = start_row + n
    last_col_letter = get_column_letter(len(header))
    table = Table(displayName="tblReconciliation", ref=f"A{start_row}:{last_col_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    for c_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18
    ws.freeze_panes = f"A{start_row + 1}"


def build_monthly_mis_sheet(wb, invoices: pd.DataFrame) -> None:
    ws = wb.create_sheet("Monthly MIS")
    ws["A1"] = "Monthly MIS - billed revenue and invoice counts via SUMIFS/COUNTIFS on Invoices"
    ws["A1"].font = TITLE_FONT

    months = sorted(pd.to_datetime(invoices["billing_period_start"]).dt.strftime("%Y-%m").unique().tolist())
    header = [
        "billing_period_start_month", "invoice_count", "expected_revenue", "billed_revenue",
        "revenue_variance", "period_start (helper)", "period_end (helper)",
    ]
    start_row = 3
    for c_idx, col in enumerate(header, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    n_inv = len(invoices)
    for i, month in enumerate(months, start=1):
        r = start_row + i
        ws.cell(row=r, column=1, value=month)
        # Helper columns build real date bounds from the "YYYY-MM" label,
        # so SUMIFS/COUNTIFS compare date-to-date rather than date-to-text
        # (text comparison against a date-typed column is locale-fragile).
        ws.cell(row=r, column=6, value=f'=DATE(VALUE(LEFT(A{r},4)),VALUE(MID(A{r},6,2)),1)')
        ws.cell(row=r, column=7, value=f'=EDATE(F{r},1)')
        ws.cell(row=r, column=6).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=7).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=(
            f'=COUNTIFS(tblInvoices[billing_period_start],">="&F{r},'
            f'tblInvoices[billing_period_start],"<"&G{r})'
        ))
        ws.cell(row=r, column=3, value=(
            f'=SUMIFS(tblInvoices[expected_fee],tblInvoices[billing_period_start],">="&F{r},'
            f'tblInvoices[billing_period_start],"<"&G{r})'
        ))
        ws.cell(row=r, column=4, value=(
            f'=SUMIFS(tblInvoices[billed_fee],tblInvoices[billing_period_start],">="&F{r},'
            f'tblInvoices[billing_period_start],"<"&G{r})'
        ))
        ws.cell(row=r, column=5, value=f"=D{r}-C{r}")

    for c_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 22


def build_kpi_dashboard_sheet(wb) -> None:
    ws = wb.create_sheet("KPI Dashboard", 0)
    ws["A1"] = "Finance Operations KPI Dashboard"
    ws["A1"].font = Font(bold=True, size=18, color="1F3864")
    ws["A2"] = "Sample workbook (40 stratified merchants) - live formulas, recalculates if source sheets change"
    ws["A2"].font = Font(italic=True, size=10)

    kpis = [
        ("Total Expected Revenue", "=SUM(tblInvoices[expected_fee])"),
        ("Total Billed Revenue", "=SUM(tblInvoices[billed_fee])"),
        ("Billing Accuracy %", '=ROUND(100*COUNTIF(\'Invoice Validation\'!H4:H1000,"Correct")/COUNTA(tblInvoices[invoice_id]),2)'),
        ("Total Collected (Payments, Success)", '=SUMIFS(tblPayments[payment_amount],tblPayments[payment_status],"Success")'),
        ("Revenue Realization %", "=ROUND(100*B7/B5,2)"),
        ("Total Outstanding AR", "=SUM('AR Aging'!G5:G1000)"),
        ("Invoices Analyzed", "=COUNTA(tblInvoices[invoice_id])"),
        ("Payments Analyzed", "=COUNTA(tblPayments[payment_id])"),
        ("Merchants in Sample", "=COUNTA(tblMerchants[merchant_id])"),
    ]
    start_row = 4
    for i, (label, formula) in enumerate(kpis):
        r = start_row + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=formula)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="DCE6F1")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws["A14"] = "See other sheets for full detail: Invoices, Payments, Invoice Validation, AR Aging, Collections, Revenue Reconciliation, Monthly MIS."
    ws["A14"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A14:F14")


def main() -> None:
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)

    merchants = pd.read_csv(PROCESSED_DIR / "merchants.csv")
    sample_merchants = _select_merchants(merchants)
    merchant_ids = set(sample_merchants["merchant_id"])
    print(f"Sampled {len(sample_merchants)} merchants across segments: "
          f"{sample_merchants['merchant_segment'].value_counts().to_dict()}")

    invoices = pd.read_csv(PROCESSED_DIR / "invoices.csv")
    invoices = invoices[invoices["merchant_id"].isin(merchant_ids)].reset_index(drop=True)
    invoice_ids = set(invoices["invoice_id"])

    payments = pd.read_csv(PROCESSED_DIR / "payments.csv")
    payments = payments[payments["invoice_id"].isin(invoice_ids)].reset_index(drop=True)

    collection_activity = pd.read_csv(PROCESSED_DIR / "collection_activity.csv")
    collection_activity = collection_activity[collection_activity["merchant_id"].isin(merchant_ids)].reset_index(drop=True)

    transactions = pd.read_csv(PROCESSED_DIR / "transactions.csv")
    transactions = transactions[transactions["merchant_id"].isin(merchant_ids)]
    if len(transactions) > MAX_TRANSACTION_ROWS:
        transactions = transactions.sample(n=MAX_TRANSACTION_ROWS, random_state=RNG_SEED)
    transactions = transactions.sort_values(["merchant_id", "transaction_date"]).reset_index(drop=True)

    print(f"Invoices: {len(invoices):,} | Payments: {len(payments):,} | "
          f"Collection activity: {len(collection_activity):,} | Transactions (sampled): {len(transactions):,}")

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True  # force Excel to recompute every formula on open

    build_merchants_sheet(wb, sample_merchants)
    build_raw_transactions_sheet(wb, transactions)
    build_invoices_sheet(wb, invoices)
    build_payments_sheet(wb, payments)
    build_invoice_validation_sheet(wb, invoices)
    build_ar_aging_sheet(wb, invoices, as_of_date="2026-08-13")
    build_collections_sheet(wb, collection_activity)
    build_revenue_reconciliation_sheet(wb, invoices)
    build_monthly_mis_sheet(wb, invoices)
    build_kpi_dashboard_sheet(wb)  # inserted at index 0 so it opens first

    wb.save(OUTPUT_PATH)
    print(f"\nWorkbook written: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
